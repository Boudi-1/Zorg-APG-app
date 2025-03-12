from flask import Flask, request, render_template, send_file, redirect, url_for, jsonify, after_this_request, flash, current_app, Response, Blueprint
import os
import pdfplumber
import re
import openpyxl
import zipfile
from werkzeug.utils import secure_filename
import shutil

# Define the blueprint at the top of the file
app2_blueprint = Blueprint('app2', __name__, template_folder='.', static_folder='static')

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "facturen_in_excel"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Vaste rijnummers voor afkortingen in de Excel-template
afkortingen_rijen = {
    "AOP": 3, "ML": 4, "BPFBA": 5, "VPBW": 6, "BEXC": 7, "BPFNTANT": 8, "BPFSGB": 9,
    "OOBW": 12, "OOBWU": 13, "COVAFB": 14, "OOAFB": 15, "FYSAFB": 16, "SFBIT": 17,
    "WAGBIT": 18, "COVBIT": 19, "COVTIM": 20, "OOTIM": 21, "SWTIM": 22, "FBORAS": 23,
    "SABBU": 26, "SABW": 27
}

def extract_data_from_pdf(file_path):
    """
    Extracts date and product amounts from PDF invoices.
    Uses multiple approaches to ensure reliable extraction regardless of PDF structure.
    """
    data = {}
    
    # Initialize data with empty values to avoid returning zeros on error
    data["datum"] = ""
    for afkorting in afkortingen_rijen.keys():
        data[afkorting] = None  # Use None instead of 0 to indicate no value was found

    try:
        with pdfplumber.open(file_path) as pdf:
            # Extract text from all pages
            full_text = " ".join([page.extract_text() for page in pdf.pages if page.extract_text()])

            # 1. Extract date
            # Look for Factuurdatum
            datum_match = re.search(r"Factuurdatum:\s*(\d{2}\.\d{2}\.\d{4})", full_text)
            if datum_match:
                data["datum"] = datum_match.group(1)
            else:
                # Try alternative date format
                datum_matches = re.findall(r"\b\d{2}\.\d{2}\.\d{4}\b", full_text)
                if datum_matches:
                    data["datum"] = datum_matches[0]
                else:
                    data["datum"] = ""

            # 2. Focus on the last page for the table with products and amounts
            last_page = pdf.pages[-1]
            if last_page:
                last_page_text = last_page.extract_text() if last_page else ""

                # Debug output
                print(f"Debug - Last page text for {file_path}:")
                print(last_page_text)

                # Try multiple strategies to extract abbreviations and amounts

                # Strategy 1: Find the table section and parse it
                table_section = ""
                if "Periode van" in last_page_text and "Totaal factuur" in last_page_text:
                    table_match = re.search(r"Periode van[^\n]+\n(.*?)Totaal factuur", last_page_text, re.DOTALL)
                    if table_match:
                        table_section = table_match.group(1)

                if table_section:
                    # Debug output
                    print("Debug - Extracted table section:")
                    print(table_section)

                    lines = table_section.strip().split('\n')
                    for line in lines:
                        # Skip irrelevant lines
                        if not line.strip() or "Product" in line or "Omschrijving" in line or "Bedrag" in line:
                            continue

                        # Check if this line starts with a known abbreviation
                        for abbr in afkortingen_rijen.keys():
                            if line.strip().startswith(abbr):
                                # Look for Euro amount
                                amount_match = re.search(r'€\s*([\d\.,]+)', line)
                                if amount_match:
                                    amount_str = amount_match.group(1)
                                    try:
                                        amount = float(amount_str.replace('.', '').replace(',', '.'))
                                        data[abbr] = amount
                                    except (ValueError, IndexError):
                                        continue
                                break  # Found a match for this line, move to next line

                # Strategy 2: Direct pattern matching for each abbreviation
                for abbr in afkortingen_rijen.keys():
                    if abbr not in data:  # Only look for abbreviations we haven't found yet
                        patterns = [
                            rf"{abbr}\s+[^\n€]*€\s*([\d\.,]+)",  # Standard pattern with Euro symbol
                            rf"{abbr}\s+[^\n]*?([\d\.]+,[\d]+)",  # Pattern with number containing comma
                            rf"{abbr}.*?([\d\.]+,[\d]+)"  # Even more flexible pattern
                        ]

                        for pattern in patterns:
                            matches = re.findall(pattern, last_page_text)
                            if matches:
                                try:
                                    amount = float(matches[0].replace('.', '').replace(',', '.'))
                                    data[abbr] = amount
                                    break  # Found a match with this pattern, try next abbreviation
                                except (ValueError, IndexError):
                                    continue

                # Strategy 3: Extract tables directly (if supported by pdfplumber)
                if any(abbr not in data for abbr in afkortingen_rijen.keys()):
                    try:
                        tables = last_page.extract_tables()
                        for table in tables:
                            for row in table:
                                if len(row) >= 2 and row[0]:
                                    product = row[0].strip()
                                    if product in afkortingen_rijen and product not in data:
                                        # Try to find amount in this row
                                        for cell in row[1:]:
                                            if cell and ('€' in cell or ',' in cell):
                                                # This might be an amount
                                                amount_str = cell.replace('€', '').strip()
                                                try:
                                                    amount = float(amount_str.replace('.', '').replace(',', '.'))
                                                    data[product] = amount
                                                    break
                                                except (ValueError, IndexError):
                                                    continue
                    except Exception as e:
                        print(f"Table extraction error: {e}")

                # Strategy 4: Last resort - flexible pattern matching with proximity detection
                if any(abbr not in data for abbr in afkortingen_rijen.keys()):
                    # Look for each abbreviation and find the nearest number
                    for abbr in afkortingen_rijen.keys():
                        if abbr not in data:  # Skip if already found
                            abbr_pos = last_page_text.find(abbr)
                            if abbr_pos >= 0:
                                # Look for numbers within 200 characters after the abbreviation
                                search_text = last_page_text[abbr_pos:abbr_pos+200]
                                # Find any number pattern that could be an amount
                                amount_matches = re.findall(r'€\s*([\d\.,]+)|\b([\d\.]+,[\d]+)\b', search_text)
                                if amount_matches:
                                    for match in amount_matches:
                                        # Try both capture groups
                                        for amount_str in match:
                                            if amount_str:
                                                try:
                                                    amount = float(amount_str.replace('.', '').replace(',', '.'))
                                                    data[abbr] = amount
                                                    break
                                                except (ValueError, IndexError):
                                                    continue
                                        if abbr in data:
                                            break  # Successfully found an amount

        # Debug: Print the extracted data
        print("Debug - Extracted data:")
        print(data)
    except Exception as e:
        print(f"Error extracting data from PDF {file_path}: {e}")
        # Don't override existing data with zeros, just return what we have
        # We already initialized data with None values at the beginning

    return data

def fill_excel_template(excel_path, pdf_files, output_filename):
    """Vult de Excel-template met de geëxtraheerde data."""
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb.active

    # Verzamel data van alle PDF's
    factuur_data_list = [extract_data_from_pdf(file) for file in pdf_files]

    for col_index, factuur_data in enumerate(factuur_data_list, start=2):  # Begin in kolom B
        col_letter = openpyxl.utils.get_column_letter(col_index)

        # Vul datum in rij 2
        ws[f"{col_letter}2"] = factuur_data.get("datum", "")

        # Vul bedragen in voor alle gedefinieerde afkortingen
        for afkorting, rij in afkortingen_rijen.items():
            if afkorting in factuur_data and factuur_data[afkorting] is not None:
                ws[f"{col_letter}{rij}"] = factuur_data[afkorting]
            # If value is None, leave the cell as is (don't set it to 0)

    output_path = os.path.join(OUTPUT_FOLDER, output_filename)
    wb.save(output_path)
    return output_path

@app2_blueprint.route('/')
def index():
    """Render de uploadpagina."""
    try:
        # Probeer het index.html bestand direct te renderen
        if os.path.exists('index.html'):
            with open('index.html', 'r', encoding='utf-8') as f:
                content = f.read()
                return content, 200, {'Content-Type': 'text/html'}
        elif os.path.exists(os.path.join('app2', 'index.html')):
            with open(os.path.join('app2', 'index.html'), 'r', encoding='utf-8') as f:
                content = f.read()
                return content, 200, {'Content-Type': 'text/html'}
        else:
            # Als het bestand niet gevonden wordt, geef een foutmelding
            return "Index.html niet gevonden", 404
    except Exception as e:
        print(f"Error in app2 index function: {e}")
        return f"Error: {str(e)}", 500

@app2_blueprint.route('/upload', methods=['POST'])
def upload_files():
    """Verwerk meerdere folders en hun bestanden."""
    try:
        uploaded_files = request.files.getlist("files")
        print(f"Ontvangen bestanden: {len(uploaded_files)}")

        if not uploaded_files or len(uploaded_files) == 0 or uploaded_files[0].filename == '':
            print("Geen bestanden ontvangen of lege bestandsnaam")
            return jsonify({"error": "Geen bestanden ontvangen."}), 400

        # Debug informatie
        for file in uploaded_files:
            print(f"Bestand: {file.filename}, type: {file.content_type}")

        # Verwerk de bestanden
        folders = {}
        for file in uploaded_files:
            # Controleer of het bestand een geldige naam heeft
            if not file.filename or file.filename == '':
                continue

            # Bepaal de mapnaam (gebruik de bestandsnaam als er geen mapstructuur is)
            parts = file.filename.split('/')
            if len(parts) > 1:
                folder_name = parts[0]  # Pak de mapnaam
            else:
                # Als er geen mapstructuur is, gebruik de bestandsnaam zonder extensie
                folder_name = os.path.splitext(file.filename)[0]

            if folder_name not in folders:
                folders[folder_name] = {"excel": None, "pdf_files": []}

            if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
                folders[folder_name]["excel"] = file
            elif file.filename.endswith(".pdf"):
                folders[folder_name]["pdf_files"].append(file)

        # Debug informatie
        print(f"Verwerkte mappen: {list(folders.keys())}")

        # Controleer of er mappen zijn om te verwerken
        if not folders:
            print("Geen geldige mappen gevonden om te verwerken")
            return jsonify({"error": "Geen geldige bestanden ontvangen."}), 400

        # Verwerk de bestanden per map
        processed_files = []
        for folder_name, files in folders.items():
            # Controleer of de map zowel Excel als PDF bestanden bevat
            if files["excel"] and files["pdf_files"]:
                print(f"Verwerken van map: {folder_name}")

                # Maak de map aan als deze nog niet bestaat
                folder_path = os.path.join(UPLOAD_FOLDER, secure_filename(folder_name))
                os.makedirs(folder_path, exist_ok=True)

                # Sla het Excel bestand op
                excel_path = os.path.join(folder_path, secure_filename(files["excel"].filename))
                files["excel"].save(excel_path)
                print(f"Excel bestand opgeslagen: {excel_path}")

                # Sla de PDF bestanden op
                pdf_paths = []
                for pdf_file in files["pdf_files"]:
                    pdf_path = os.path.join(folder_path, secure_filename(pdf_file.filename))
                    pdf_file.save(pdf_path)
                    pdf_paths.append(pdf_path)
                    print(f"PDF bestand opgeslagen: {pdf_path}")

                # Vul de Excel template in met de data uit de PDF bestanden
                output_filename = f"{folder_name}_verwerkt.xlsx"
                try:
                    output_path = fill_excel_template(excel_path, pdf_paths, output_filename)
                    processed_files.append(output_path)
                    print(f"Excel template ingevuld: {output_path}")
                except Exception as e:
                    print(f"Fout bij het invullen van Excel template: {e}")
            else:
                print(f"Map {folder_name} bevat geen Excel bestand of geen PDF bestanden")

        # Controleer of er bestanden zijn verwerkt
        if not processed_files:
            print("Geen bestanden verwerkt")
            return jsonify({"error": "Geen bestanden konden worden verwerkt."}), 400

        # Maak een ZIP-bestand met alle ingevulde Excel-bestanden
        try:
            zip_path = os.path.join(OUTPUT_FOLDER, "facturen_ingevuld.zip")
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for file in processed_files:
                    zipf.write(file, os.path.basename(file))
            print(f"ZIP bestand aangemaakt: {zip_path}")
        except Exception as e:
            print(f"Fout bij het aanmaken van ZIP bestand: {e}")

        # Bepaal de juiste redirect URL
        try:
            # Probeer eerst de success route in de huidige blueprint
            success_url = url_for('app2.success', folders=",".join(folders.keys()))
            print(f"Redirect naar: {success_url}")
            return redirect(success_url)
        except Exception as e:
            print(f"Error redirecting to app2.success: {e}")
            try:
                # Probeer dan de success route in de hoofdapplicatie
                success_url = '/app2/success?folders=' + ",".join(folders.keys())
                print(f"Redirect naar: {success_url}")
                return redirect(success_url)
            except Exception as e:
                print(f"Error redirecting to /app2/success: {e}")
                # Als laatste optie, redirect naar de success.html pagina
                return redirect('/app2/success.html?folders=' + ",".join(folders.keys()))
    except Exception as e:
        print(f"Error in upload_files: {e}")
        return jsonify({"error": str(e)}), 500

@app2_blueprint.route('/success')
def success():
    """Toon de pagina met de verwerkte mappen en downloadknop."""
    try:
        folders = request.args.get("folders", "")
        folder_list = folders.split(",") if folders else []

        # Gebruik Flask's render_template functie om de template correct te renderen
        try:
            from flask import render_template

            # Probeer verschillende locaties voor de success.html template
            if os.path.exists('app2/success.html'):
                # Lees de template en render deze met de juiste context
                with open('app2/success.html', 'r', encoding='utf-8') as f:
                    template_content = f.read()

                from flask import render_template_string
                return render_template_string(template_content, folders=folder_list)
            else:
                # Als geen template gevonden wordt, toon een eenvoudige HTML pagina
                html = """
                <!DOCTYPE html>
                <html lang="nl">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Verwerking Geslaagd</title>
                    <script src="https://cdn.tailwindcss.com"></script>
                </head>
                <body class="bg-gradient-to-br from-green-50 to-green-100 min-h-screen flex items-center justify-center p-4">
                    <div class="w-full max-w-xl bg-white rounded-3xl shadow-2xl p-8">
                        <div class="text-center mb-6">
                            <div class="flex justify-center mb-4">
                                <svg class="w-16 h-16 text-green-600" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z"></path>
                                </svg>
                            </div>
                            <h1 class="text-3xl font-bold mb-2 text-green-800">Verwerking Geslaagd!</h1>
                            <p class="text-gray-600 text-lg">De volgende mappen zijn succesvol verwerkt:</p>
                        </div>
                """

                # Voeg de lijst met mappen toe
                if folder_list:
                    html += """
                        <div class="bg-green-50 p-4 rounded-lg mb-6 max-h-48 overflow-y-auto">
                            <ul class="space-y-2">
                    """

                    for folder in folder_list:
                        html += f"""
                                <li class="flex items-center bg-white p-3 rounded-md shadow-sm">
                                    <svg class="w-5 h-5 text-green-500 mr-3" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 7v10a2 2 0 002 2h14a2 2 0 002-2V9a2 2 0 00-2-2h-6l-2-2H5a2 2 0 00-2 2z"></path>
                                    </svg>
                                    <span class="text-gray-700 font-medium">{folder}</span>
                                </li>
                        """

                    html += """
                            </ul>
                        </div>
                    """
                else:
                    html += """
                        <p class="text-red-600 bg-red-50 p-4 rounded-lg text-center">Geen mappen verwerkt.</p>
                    """

                # Voeg de knoppen toe
                html += """
                        <div class="flex flex-col space-y-4">
                            <a
                                href="/app2/download"
                                class="w-full bg-green-600 hover:bg-green-700 text-white font-bold py-3 rounded-lg transition-all duration-300 transform hover:scale-105 flex items-center justify-center text-center"
                            >
                                <svg class="w-6 h-6 mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"></path>
                                </svg>
                                Download Resultaten
                            </a>

                            <a
                                href="/app2/"
                                class="w-full text-center text-green-700 hover:text-green-900 font-semibold py-3 rounded-lg transition-colors bg-green-100 hover:bg-green-200"
                            >
                                Terug naar uploadpagina
                            </a>

                            <a
                                href="/terug-naar-dashboard"
                                class="w-full text-center text-blue-700 hover:text-blue-900 font-semibold py-3 rounded-lg transition-colors bg-blue-100 hover:bg-blue-200"
                            >
                                Terug naar Dashboard
                            </a>
                        </div>
                    </div>
                </body>
                </html>
                """

                return html
        except Exception as e:
            print(f"Error rendering success template: {e}")
            # Fallback naar een eenvoudige HTML pagina
            return f"""
            <html>
            <body>
                <h1>Verwerking Voltooid</h1>
                <p>De volgende mappen zijn verwerkt: {folder_list}</p>
                <a href="/app2/download">Download Verwerkte Excel</a>
                <a href="/">Terug naar Dashboard</a>
            </body>
            </html>
            """, 200, {'Content-Type': 'text/html'}
    except Exception as e:
        print(f"Error in success function: {e}")
        return f"Er is een fout opgetreden: {str(e)}", 500

@app2_blueprint.route('/download')
def download():
    """Laat de gebruiker het ZIP-bestand downloaden."""
    try:
        # Zoek het ZIP bestand
        zip_path = os.path.join(OUTPUT_FOLDER, "facturen_ingevuld.zip")

        # Controleer of het ZIP bestand bestaat
        if not os.path.exists(zip_path):
            # Probeer alternatieve locaties
            alternative_paths = [
                os.path.join('facturen_in_excel', "facturen_ingevuld.zip"),
                os.path.join('app2', 'facturen_in_excel', "facturen_ingevuld.zip")
            ]

            for alt_path in alternative_paths:
                if os.path.exists(alt_path):
                    zip_path = alt_path
                    break
            else:
                return "Fout: Het bestand bestaat niet. Verwerk eerst facturen.", 404

        @after_this_request
        def cleanup(response):
            """Leegt de uploads en output folders na verwerking, nadat de response volledig is verzonden."""
            try:
                # Maak een kopie van het ZIP-bestand buiten de te legen mappen
                if os.path.exists(zip_path):
                    backup_zip = os.path.join(os.path.dirname(OUTPUT_FOLDER), "facturen_ingevuld_backup.zip")
                    shutil.copy2(zip_path, backup_zip)

                    # Leeg de mappen
                    if os.path.exists(UPLOAD_FOLDER):
                        shutil.rmtree(UPLOAD_FOLDER, ignore_errors=True)
                    if os.path.exists(OUTPUT_FOLDER):
                        shutil.rmtree(OUTPUT_FOLDER, ignore_errors=True)

                    # Maak de mappen opnieuw aan
                    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

                    # Zet het ZIP-bestand terug
                    shutil.copy2(backup_zip, zip_path)
                    os.remove(backup_zip)  # Verwijder de backup

                    print("Folders zijn succesvol geleegd.")
            except Exception as e:
                print(f"Fout bij het legen van de folders: {e}")

            return response

        # Stuur het ZIP-bestand naar de gebruiker
        return send_file(zip_path, as_attachment=True, download_name="facturen_ingevuld.zip")
    except Exception as e:
        print(f"Error in download function: {e}")
        return f"Er is een fout opgetreden bij het downloaden: {str(e)}", 500

def register_app2(app):
    """Registreert app2 bij de hoofdapplicatie."""
    # Registreer de blueprint met een url_prefix
    app.register_blueprint(app2_blueprint, url_prefix='/app2')

    print("App2 succesvol geregistreerd")
    return True

if __name__ == "__main__":
    # Create a Flask app for standalone testing
    app = Flask(__name__)
    app.register_blueprint(app2_blueprint)
    app.run(debug=True)





